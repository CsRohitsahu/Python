import openpyxl
path="D:\\PHYTHON FULL COURSE\\pandas\\contact.xlsx"
wb=openpyxl.load_workbook(path)
sheet=wb.active
cell_obj=sheet.cell(row=1,column=1)
print(cell_obj)
row=sheet.max_row
column=sheet.max_column
print('\n value of first column ')
#show the column of sheet
for i in range (1,row+1):
    cell_obj=sheet.cell(row=i,column =1)
    print(cell_obj.value,end=" ")
#show the row of sheet  
print("\nvalue of rist row ")   
for i in range (1,column+1):
        cell_obj=sheet.cell(row=2,column =i)
        print(cell_obj.value,end=" ")
print("\printign mulitple cell\n")
cell_obj=sheet['A1':'B3']
for cell1,cell2 in cell_obj:
      print(cell1.value,cell2.value)
















# reading from multiple cells 
path="D:\\PHYTHON FULL COURSE\\pandas\\contact.xlsx"
#open the workbook 
wb=openpyxl.load_workbook(path)

  
      